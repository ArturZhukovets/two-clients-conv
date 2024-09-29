#####################################################################################################
import csv
import io
import zipfile
from collections.abc import Sequence
from datetime import datetime
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Border, PatternFill, Side
from starlette.requests import Request
from starlette.responses import StreamingResponse

from l7x.db import TextModel
from l7x.utils.datetime_utils import format_datetime_to_iso
from l7x.utils.db_utils import check_session_with_request, check_superuser_state
from l7x.utils.fastapi_utils import AppFastAPI

#####################################################################################################

async def prepare_conversation_rows(texts: Sequence[TextModel]):
    text_rows = []
    audio_for_zip = []

    headers = [
        'UUID Сообщения', 'UUID Audio', 'Создано', 'Язык', 'Распознанный текст', 'Исправленный текст'
    ]
    text_rows.append(headers)
    for text in texts:
        audio_obj = text.audio_id
        text_data_record = [
            str(text.primary_uuid),
            str(audio_obj.primary_uuid),
            format_datetime_to_iso(text.create_ts),
            text.lang_from,
            text.recognized_text,
            text.fixed_text,
        ]
        text_rows.append(text_data_record)
        audio_path = f'wav/{audio_obj.primary_uuid}.wav' if audio_obj is not None else None
        if audio_path is not None:
            audio_for_zip.append((audio_path, io.BytesIO(text.audio_id.audio_raw)))

    return text_rows, audio_for_zip

#####################################################################################################

async def _texts_data_page(request: Request) -> StreamingResponse:
    is_valid_session = await check_session_with_request(request, 'create transcribe audio error report')

    error_headers = {
        'Content-Disposition': f'attachment; filename="messages_download_error.txt"'
    }

    if not is_valid_session:
        error_file_buffer = io.BytesIO(b'Session is not valid')
        err_response = StreamingResponse(error_file_buffer, headers=error_headers)
        err_response.delete_cookie('session_uuid')
        return err_response

    if not await check_superuser_state(request):
        error_file_buffer = io.BytesIO(b'The user does not have superuser rights')
        return StreamingResponse(error_file_buffer, headers=error_headers)

    data: Final = await request.json()
    text_primary_uuid: Final = data.get('primary_uuid')
    file_extension: Final = data.get('report_format', 'xlsx')
    files_for_zip = []

    if isinstance(text_primary_uuid, list):
        primary_uuid_list = text_primary_uuid
    elif isinstance(text_primary_uuid, str):
        primary_uuid_list = [text_primary_uuid]
    else:
        primary_uuid_list = []

    texts: Sequence[TextModel] = await TextModel.objects.select_related(
        'audio_id',
    ).order_by(
        ['create_ts'],
    ).filter(
        primary_uuid__in=primary_uuid_list,
    ).all()

    if texts:
        text_rows, audio_for_zip = await prepare_conversation_rows(texts)
        report_buffer = io.BytesIO()
        if file_extension == 'xlsx':
            work_book = openpyxl.Workbook()
            work_sheet = work_book.active
            work_sheet.title = "Conversations"

            for text_row in text_rows:
                work_sheet.append(text_row)
            work_sheet.freeze_panes = 'A2'

            fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thick_border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

            for row in work_sheet.iter_rows(min_row=1, max_row=work_sheet.max_row, min_col=1, max_col=work_sheet.max_column):
                for cell in row:
                    if cell.value == ' ':
                        cell.fill = fill
                    else:
                        cell.border = thick_border

            for column_cells in work_sheet.columns:
                max_length = 0
                column = list(column_cells)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = min(len(cell.value), 25)
                        cell.alignment = align
                    except:
                        pass
                adjusted_width = (max_length + 2)
                work_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            work_book.save(report_buffer)
        elif file_extension == 'csv':
            text_buffer = io.TextIOWrapper(report_buffer, encoding='utf-8-sig', newline='')
            csv_writer = csv.writer(text_buffer)
            csv_writer.writerows(text_rows)
            text_buffer.flush()
        else:
            error_file_buffer = io.BytesIO(f'Not supported file extension: {file_extension}'.encode())
            return StreamingResponse(error_file_buffer, headers=error_headers)
        report_buffer.seek(0)

        report_file_name: Final = f'report_{datetime.utcnow().strftime("%d-%m-%Y%-H-%M-%S")}'
        files_for_zip.append((f'{report_file_name}.{file_extension}', report_buffer))

        for audio_zip_data in audio_for_zip:
            files_for_zip.append(audio_zip_data)

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for file_name, data in files_for_zip:
                zip_file.writestr(file_name, data.getvalue())

        zip_buffer.seek(0)

        headers = {
            'Content-Disposition': f'attachment; filename="{report_file_name}.zip"'
        }

        return StreamingResponse(zip_buffer, headers=headers)
    else:
        error_file_buffer = io.BytesIO(b'Conversation not found')
        return StreamingResponse(error_file_buffer, headers=error_headers)

#####################################################################################################

def text_data_page_listener_registrar(app: AppFastAPI, /) -> None:
    if app.app_settings.enable_admin_pages:
        app.post('/api/text_data')(_texts_data_page)

#####################################################################################################
