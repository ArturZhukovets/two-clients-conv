#####################################################################################################
import csv
import io
import zipfile
from collections.abc import Sequence
from datetime import datetime
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Border, PatternFill, Side
from starlette.requests import Request
from starlette.responses import StreamingResponse

from l7x.configs.settings import AppSettings
from l7x.db import ConversationModel
from l7x.utils.datetime_utils import format_datetime_to_iso
from l7x.utils.db_utils import check_session_with_request, check_superuser_state
from l7x.utils.fastapi_utils import AppFastAPI
from l7x.types.localization import TKey
from l7x.utils.lang_utils import localize as _

#####################################################################################################

def _get_common_values(conversation: ConversationModel, app_settings: AppSettings):
    values = [
        str(conversation.primary_uuid),
        format_datetime_to_iso(conversation.start_ts),
        format_datetime_to_iso(conversation.end_ts) if conversation.end_ts is not None else '',
        conversation.session_id.user_id.department_id.name,
        conversation.session_id.user_id.full_name,
        conversation.selected_lang,
    ]
    if app_settings.enable_questionnaire:
        scores = conversation.questionare if conversation.questionare is not None else {}
        values.extend(
            (
                scores.get('recommends', '-'),
                scores.get('translation_quality', '-'),
                scores.get('difficulty_of_use', '-'),
            )
        )
    return values

#####################################################################################################

def _get_headers(app_settings: AppSettings) -> Sequence[str]:
    """Return without scores headers if not 'enable_questionnaire' """
    if not app_settings.enable_questionnaire:
        return [
            _(TKey.A_DIALOG_UUID), _(TKey.A_DIALOG_START), _(TKey.A_DIALOG_END),
            _(TKey.A_DEPARTMENT), _(TKey.A_OPERATOR_NAME),
            _(TKey.A_LANGUAGE), _(TKey.A_MESSAGE_UUID), _(TKey.A_AUDIO_UUID),
            _(TKey.A_MESSAGE_TYPE), _(TKey.A_MESSAGE_START), _(TKey.A_RECOGNIZED_TEXT), _(TKey.A_TRANSLATED_TEXT),
            _(TKey.A_SOURCE_LANGUAGE), _(TKey.A_TARGET_LANGUAGE), _(TKey.A_EDITING_TIME), _(TKey.A_EDITED_TEXT),
        ]
    return [
        _(TKey.A_DIALOG_UUID), _(TKey.A_DIALOG_START), _(TKey.A_DIALOG_END),
        _(TKey.A_DEPARTMENT), _(TKey.A_OPERATOR_NAME),
        _(TKey.A_LANGUAGE), _(TKey.A_NPS_SCORE), _(TKey.A_TRANSLATION_SCORE),
        _(TKey.A_USABILITY_SCORE), _(TKey.A_MESSAGE_UUID), _(TKey.A_AUDIO_UUID),
        _(TKey.A_MESSAGE_TYPE), _(TKey.A_MESSAGE_START), _(TKey.A_RECOGNIZED_TEXT), _(TKey.A_TRANSLATED_TEXT),
        _(TKey.A_SOURCE_LANGUAGE), _(TKey.A_TARGET_LANGUAGE), _(TKey.A_EDITING_TIME), _(TKey.A_EDITED_TEXT),
    ]

#####################################################################################################

async def prepare_conversation_rows(
    conversations: Sequence[ConversationModel],
    is_download_audio: bool,
    app_settings: AppSettings,
):
    conversation_rows = []
    audio_for_zip = []

    headers = _get_headers(app_settings)

    conversation_rows.append(headers)
    for conversation in conversations:
        common_values = _get_common_values(conversation, app_settings)
        conv_record = common_values.copy()
        conv_record.extend(
            (' ' for _ in range(len(headers) - len(conv_record)))
        )

        conversation_rows.append(conv_record)
        for text in conversation.texts:
            audio_obj = text.audio_id
            text_data_record = common_values.copy()
            text_data_record.extend(
                (
                    str(text.primary_uuid),
                    str(audio_obj.primary_uuid) if audio_obj is not None else '-',
                    text.type,
                    format_datetime_to_iso(text.create_ts),
                    text.recognized_text,
                    text.translated_text,
                    text.lang_from,
                    text.lang_to,
                    format_datetime_to_iso(text.edit_ts) if text.edit_ts is not None else '',
                    text.fixed_text,
                )
            )
            conversation_rows.append(text_data_record)
            audio_path = f'wav/{conversation.primary_uuid}/{audio_obj.primary_uuid}.wav' if audio_obj is not None else None
            if audio_path is not None and is_download_audio:
                audio_for_zip.append((audio_path, io.BytesIO(text.audio_id.audio_raw)))

    return conversation_rows, audio_for_zip

#####################################################################################################

async def _conversation_data_page(request: Request) -> StreamingResponse:
    app_settings: Final = request.app.app_settings
    is_valid_session = await check_session_with_request(request, 'create conversations data report')
    error_headers = {
        'Content-Disposition': f'attachment; filename="conv_download_error.txt"'
    }
    if not is_valid_session:
        error_file_buffer = io.BytesIO(b'Session is not valid')
        err_response = StreamingResponse(error_file_buffer, headers=error_headers)
        err_response.delete_cookie('session_uuid')
        return err_response

    if not await check_superuser_state(request):
        error_file_buffer = io.BytesIO(b'The user does not have superuser rights')
        return StreamingResponse(error_file_buffer, headers=error_headers)

    data: Final = await request.json()
    primary_uuid: Final = data.get('primary_uuid')
    file_extension: Final = data.get('report_format', 'xlsx')
    is_download_audio: Final = data.get('is_download_audio', True)
    files_for_zip = []

    if isinstance(primary_uuid, list):
        primary_uuid_list = primary_uuid
    elif isinstance(primary_uuid, str):
        primary_uuid_list = [primary_uuid]
    else:
        primary_uuid_list = []

    conversations: Sequence[ConversationModel] = await ConversationModel.objects.select_related(
        'session_id__user_id__department_id',
    ).prefetch_related(
        'texts__audio_id',
    ).order_by(
        ['start_ts', 'texts__create_ts'],
    ).filter(
        primary_uuid__in=primary_uuid_list
    ).all()

    if conversations:
        conversation_rows, audio_for_zip = await prepare_conversation_rows(conversations, is_download_audio, app_settings)
        report_buffer = io.BytesIO()
        if file_extension == 'xlsx':
            work_book = openpyxl.Workbook()
            work_sheet = work_book.active
            work_sheet.title = "Conversations"

            for conversation_row in conversation_rows:
                work_sheet.append(conversation_row)
            work_sheet.freeze_panes = 'A2'

            fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thick_border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

            for row in work_sheet.iter_rows(min_row=1, max_row=work_sheet.max_row, min_col=1, max_col=work_sheet.max_column):
                for cell in row:
                    if cell.value == ' ':
                        cell.fill = fill
                    else:
                        cell.border = thick_border

            for column_cells in work_sheet.columns:
                max_length = 0
                column = list(column_cells)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = min(len(cell.value), 25)
                        cell.alignment = align
                    except:
                        pass
                adjusted_width = (max_length + 2)
                work_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            work_book.save(report_buffer)
        elif file_extension == 'csv':
            text_buffer = io.TextIOWrapper(report_buffer, encoding='utf-8-sig', newline='')
            csv_writer = csv.writer(text_buffer)
            csv_writer.writerows(conversation_rows)
            text_buffer.flush()
        else:
            error_file_buffer = io.BytesIO(f'Not supported file extension: {file_extension}'.encode())
            return StreamingResponse(error_file_buffer, headers=error_headers)
        report_buffer.seek(0)

        report_file_name: Final = f'report_{datetime.utcnow().strftime("%d-%m-%Y%-H-%M-%S")}'
        files_for_zip.append((f'{report_file_name}.{file_extension}', report_buffer))

        for audio_zip_data in audio_for_zip:
            files_for_zip.append(audio_zip_data)

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for file_name, data in files_for_zip:
                zip_file.writestr(file_name, data.getvalue())

        zip_buffer.seek(0)

        headers = {
            'Content-Disposition': f'attachment; filename="{report_file_name}.zip"'
        }

        return StreamingResponse(zip_buffer, headers=headers)
    else:
        error_file_buffer = io.BytesIO(b'Conversation not found')
        return StreamingResponse(error_file_buffer, headers=error_headers)

#####################################################################################################

def conversation_data_page_listener_registrar(app: AppFastAPI, /) -> None:
    if app.app_settings.enable_admin_pages:
        app.post('/api/conversation_data')(_conversation_data_page)

#####################################################################################################
